from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[1]
PLANNING_DIR = REPO_ROOT / "planning"
MANIFEST_PATH = PLANNING_DIR / "tracker_manifest.json"
EXPORTS_DIR = PLANNING_DIR / "exports"
WORKBOOKS_DIR = PLANNING_DIR / "workbooks"
GRAPH_PATH = REPO_ROOT / ".archex" / "graph.json"
METRICS_PATH = REPO_ROOT / ".archex" / "metrics.json"
IGNORED_PARTS = {".git", ".venv", "__pycache__", ".pytest_cache", ".mypy_cache", ".ruff_cache"}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def add_sheet_from_rows(workbook: Workbook, title: str, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(fieldnames)
    style_header(worksheet[1])
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
    autosize_worksheet(worksheet)


def relative_path(path: Path) -> str:
    return path.relative_to(REPO_ROOT).as_posix()


def is_ignored(path: Path) -> bool:
    return any(part in IGNORED_PARTS for part in path.parts)


def build_artifact_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for artifact in manifest["artifacts"]:
        target = REPO_ROOT / artifact["path"]
        rows.append(
            {
                "artifact_id": artifact["artifact_id"],
                "name": artifact["name"],
                "discipline": artifact["discipline"],
                "path": artifact["path"],
                "kind": artifact["kind"],
                "exists": target.exists(),
                "purpose": artifact["purpose"],
            }
        )
    return rows


def build_workstream_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for workstream in manifest["workstreams"]:
        anchor = REPO_ROOT / workstream["anchor_path"]
        evidence = REPO_ROOT / workstream["evidence_anchor"]
        rows.append(
            {
                "workstream_id": workstream["workstream_id"],
                "name": workstream["name"],
                "discipline": workstream["discipline"],
                "anchor_path": workstream["anchor_path"],
                "anchor_exists": anchor.exists(),
                "evidence_anchor": workstream["evidence_anchor"],
                "evidence_exists": evidence.exists(),
                "goal": workstream["goal"],
            }
        )
    return rows


def build_software_sequence_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return list(manifest.get("software_sequence", []))


def build_manual_stream_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for stream in manifest.get("manual_streams", []):
        rows.append(
            {
                "stream_id": stream["stream_id"],
                "name": stream["name"],
                "focus": stream["focus"],
                "status": "manual",
                "owner": "",
                "next_step": "",
                "blocking_risk": "",
                "evidence_link": "",
                "prompt_1": stream["prompt_1"],
                "prompt_2": stream["prompt_2"],
                "prompt_3": stream["prompt_3"],
            }
        )
    return rows


def build_manual_detail_rows(manifest: dict[str, Any], stream_id: str) -> list[dict[str, Any]]:
    stream = next((item for item in manifest.get("manual_streams", []) if item["stream_id"] == stream_id), None)
    if stream is None:
        return []

    return [
        {
            "item_id": f"{stream_id}-001",
            "stream": stream["name"],
            "focus_area": stream["focus"],
            "status": "not-started",
            "owner": "",
            "next_step": "",
            "blocking_risk": "",
            "evidence_link": "",
            "prompt_1": stream["prompt_1"],
            "prompt_2": stream["prompt_2"],
            "prompt_3": stream["prompt_3"],
            "notes": "",
        }
    ]


def build_milestone_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return list(manifest["milestones"])


def build_module_health_rows() -> list[dict[str, Any]]:
    if not GRAPH_PATH.exists() or not METRICS_PATH.exists():
        return []

    graph = load_json(GRAPH_PATH)
    metrics = load_json(METRICS_PATH)
    module_nodes = [node for node in graph.get("nodes", []) if node.get("kind") == "module"]

    coupling = metrics.get("coupling", {})
    cohesion = metrics.get("cohesion", {})
    rows: list[dict[str, Any]] = []

    for node in sorted(module_nodes, key=lambda item: item.get("id", "")):
        module_id = node["id"]
        coupling_info = coupling.get(module_id, {})
        cohesion_info = cohesion.get(module_id, {})
        rows.append(
            {
                "module_id": module_id,
                "file": node.get("file", ""),
                "ca": coupling_info.get("ca", ""),
                "ce": coupling_info.get("ce", ""),
                "total_coupling": coupling_info.get("total_coupling", ""),
                "lcom4": cohesion_info.get("lcom4", ""),
                "cohesion_rating": cohesion_info.get("cohesion_rating", ""),
                "components": cohesion_info.get("components", ""),
                "symbol_count": cohesion_info.get("symbol_count", ""),
            }
        )

    return rows


def count_files(pattern: str) -> int:
    return sum(1 for path in REPO_ROOT.rglob(pattern) if path.is_file() and not is_ignored(path))


def build_snapshot_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    root_dirs = [
        path for path in REPO_ROOT.iterdir() if path.is_dir() and not path.name.startswith(".")
    ]
    rows = [
        {"metric": "generated_at_utc", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "artifact_count", "value": len(manifest["artifacts"])},
        {"metric": "workstream_count", "value": len(manifest["workstreams"])},
        {"metric": "milestone_count", "value": len(manifest["milestones"])},
        {"metric": "top_level_directory_count", "value": len(root_dirs)},
        {"metric": "python_file_count", "value": count_files("*.py")},
        {"metric": "world_file_count", "value": count_files("*.wbt")},
        {"metric": "proto_file_count", "value": count_files("*.proto")},
        {"metric": "architecture_md_present", "value": (REPO_ROOT / "architecture.md").exists()},
        {"metric": "archex_metrics_present", "value": METRICS_PATH.exists()},
        {"metric": "archex_graph_present", "value": GRAPH_PATH.exists()},
        {"metric": "ci_pipeline_present", "value": (REPO_ROOT / ".github" / "workflows" / "ci.yml").exists()},
    ]
    return rows


def build_anchor_inventory() -> list[dict[str, Any]]:
    anchor_paths = [
        REPO_ROOT / "controllers",
        REPO_ROOT / "core",
        REPO_ROOT / "robots",
        REPO_ROOT / "worlds",
        REPO_ROOT / "protos",
        REPO_ROOT / ".archex",
    ]
    rows: list[dict[str, Any]] = []
    for anchor in anchor_paths:
        file_count = (
            sum(1 for path in anchor.rglob("*") if path.is_file() and not is_ignored(path))
            if anchor.exists()
            else 0
        )
        rows.append(
            {
                "anchor": relative_path(anchor),
                "exists": anchor.exists(),
                "file_count": file_count,
            }
        )
    return rows


def build_overview_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    workbook_info = manifest.get("workbook", {})
    return [
        {
            "section": "Purpose",
            "value": workbook_info.get(
                "intent",
                "One workbook that keeps software sequencing structured and non-software workstreams manually guided.",
            ),
        },
        {
            "section": "How to use",
            "value": "Drive execution from the Software Sequence sheet. Fill Mechanical, Electronics, and Manufacturing from the Manual Streams sheet by hand.",
        },
        {
            "section": "Reality link",
            "value": "Use Repo Artifacts, Snapshot, and Module Health as the live project mirror sourced from TerrestraX and ArchEx.",
        },
        {
            "section": "Generated at",
            "value": datetime.now(timezone.utc).isoformat(),
        },
    ]


def build_workbook(manifest: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    add_sheet_from_rows(
        workbook,
        "Start Here",
        build_overview_rows(manifest),
        ["section", "value"],
    )
    add_sheet_from_rows(
        workbook,
        "Software Sequence",
        build_software_sequence_rows(manifest),
        [
            "step_id",
            "sequence",
            "title",
            "objective",
            "entry_criteria",
            "done_criteria",
            "primary_anchor",
            "supporting_anchor",
            "validation",
            "depends_on",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Manual Streams",
        build_manual_stream_rows(manifest),
        [
            "stream_id",
            "name",
            "focus",
            "status",
            "owner",
            "next_step",
            "blocking_risk",
            "evidence_link",
            "prompt_1",
            "prompt_2",
            "prompt_3",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Mechanical",
        build_manual_detail_rows(manifest, "MAN-MECH"),
        [
            "item_id",
            "stream",
            "focus_area",
            "status",
            "owner",
            "next_step",
            "blocking_risk",
            "evidence_link",
            "prompt_1",
            "prompt_2",
            "prompt_3",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Electronics",
        build_manual_detail_rows(manifest, "MAN-ELEC"),
        [
            "item_id",
            "stream",
            "focus_area",
            "status",
            "owner",
            "next_step",
            "blocking_risk",
            "evidence_link",
            "prompt_1",
            "prompt_2",
            "prompt_3",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Manufacturing",
        build_manual_detail_rows(manifest, "MAN-MFG"),
        [
            "item_id",
            "stream",
            "focus_area",
            "status",
            "owner",
            "next_step",
            "blocking_risk",
            "evidence_link",
            "prompt_1",
            "prompt_2",
            "prompt_3",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Validation",
        load_csv_rows(PLANNING_DIR / "templates" / "validation_matrix.csv"),
        [
            "validation_id",
            "linked_item_id",
            "discipline",
            "validation_type",
            "target_artifact",
            "scenario_or_world",
            "acceptance_criteria",
            "evidence_location",
            "status",
            "owner",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Risks",
        load_csv_rows(PLANNING_DIR / "templates" / "risk_register.csv"),
        [
            "risk_id",
            "title",
            "category",
            "severity",
            "likelihood",
            "status",
            "owner",
            "linked_item_id",
            "trigger",
            "mitigation",
            "next_review",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "BOM",
        load_csv_rows(PLANNING_DIR / "templates" / "bom.csv"),
        [
            "bom_id",
            "part_name",
            "discipline",
            "category",
            "qty",
            "unit",
            "status",
            "source_or_vendor",
            "linked_workstream",
            "linked_milestone",
            "owner",
            "notes",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Repo Artifacts",
        build_artifact_rows(manifest),
        ["artifact_id", "name", "discipline", "path", "kind", "exists", "purpose"],
    )
    add_sheet_from_rows(
        workbook,
        "Module Health",
        build_module_health_rows(),
        [
            "module_id",
            "file",
            "ca",
            "ce",
            "total_coupling",
            "lcom4",
            "cohesion_rating",
            "components",
            "symbol_count",
        ],
    )
    add_sheet_from_rows(
        workbook,
        "Snapshot",
        build_snapshot_rows(manifest),
        ["metric", "value"],
    )

    workbook_path = WORKBOOKS_DIR / manifest.get("workbook", {}).get("name", "TerrestraX_Tracker.xlsx")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return workbook_path


def load_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    manifest = load_json(MANIFEST_PATH)
    workbook_path = build_workbook(manifest)

    write_csv(
        EXPORTS_DIR / "repo_artifacts.csv",
        build_artifact_rows(manifest),
        ["artifact_id", "name", "discipline", "path", "kind", "exists", "purpose"],
    )
    write_csv(
        EXPORTS_DIR / "workstreams.csv",
        build_workstream_rows(manifest),
        [
            "workstream_id",
            "name",
            "discipline",
            "anchor_path",
            "anchor_exists",
            "evidence_anchor",
            "evidence_exists",
            "goal",
        ],
    )
    write_csv(
        EXPORTS_DIR / "software_sequence.csv",
        build_software_sequence_rows(manifest),
        [
            "step_id",
            "sequence",
            "title",
            "objective",
            "entry_criteria",
            "done_criteria",
            "primary_anchor",
            "supporting_anchor",
            "validation",
            "depends_on",
        ],
    )
    write_csv(
        EXPORTS_DIR / "manual_streams.csv",
        build_manual_stream_rows(manifest),
        [
            "stream_id",
            "name",
            "focus",
            "status",
            "owner",
            "next_step",
            "blocking_risk",
            "evidence_link",
            "prompt_1",
            "prompt_2",
            "prompt_3",
        ],
    )
    write_csv(
        EXPORTS_DIR / "milestones.csv",
        build_milestone_rows(manifest),
        ["milestone_id", "sequence", "name", "focus", "evidence"],
    )
    write_csv(
        EXPORTS_DIR / "module_health.csv",
        build_module_health_rows(),
        [
            "module_id",
            "file",
            "ca",
            "ce",
            "total_coupling",
            "lcom4",
            "cohesion_rating",
            "components",
            "symbol_count",
        ],
    )
    write_csv(
        EXPORTS_DIR / "project_snapshot.csv",
        build_snapshot_rows(manifest),
        ["metric", "value"],
    )
    write_csv(
        EXPORTS_DIR / "anchor_inventory.csv",
        build_anchor_inventory(),
        ["anchor", "exists", "file_count"],
    )

    print(f"Generated tracker workbook at {relative_path(workbook_path)}")
    print(f"Generated tracker exports in {relative_path(EXPORTS_DIR)}")


if __name__ == "__main__":
    main()
